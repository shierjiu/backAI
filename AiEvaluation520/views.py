# AiEvaluationy/views.py
import copy
import datetime
import logging
import openpyxl
from django.db import transaction
from django.http import JsonResponse
from rest_framework import status
from rest_framework.parsers import *
from AiServer.utils import GenericAIService
from Public.api import *
from .serializers import *
import threading
import re
import logging
from rest_framework.views import APIView
from rest_framework.response import Response
from AiServer.models import AiServerAgent
from AiServer.serializers import AiServerAgentSerializer
from AiServer.utils import AIAgentServer
from .models import *
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# 评估器
class DatasetEvaluateByAgentsView(APIView):
    def post(self, request):
        dataset_id = request.data.get('dataset_id')
        agent = request.data.get('agent')
        if not dataset_id or not agent:
            return Response({"error": "缺少 dataset_id 或 agent"}, status=400)
        if not isinstance(agent, list):
            return Response({"error": "agent 必须是列表"}, status=400)
        # 1. 验证数据集
        try:
            dataset = AiEvaluationDataset.objects.get(pk=dataset_id)
        except AiEvaluationDataset.DoesNotExist:
            return Response({"error": "数据集不存在"}, status=404)
        # 2. 验证智能体
        agents_qs = AiServerAgent.objects.filter(id__in=agent)
        if agents_qs.count() != len(agent):
            return Response({"error": "所选智能体中有一个或多个不存在"}, status=404)
        # 3. 验证条目一致性 & 取 entity ID 和 name
        entries = AiEvaluationDatasetEntry.objects.filter(dataset_id=dataset_id)
        if not entries.exists():
            return Response({"error": "数据集没有可评测的条目"}, status=400)
        distinct_entity_ids = entries.values_list('entity', flat=True).distinct()
        if distinct_entity_ids.count() > 1:
            return Response({"error": "数据集条目评测对象不一致，请统一后再评测"}, status=400)
        # 先从 distinct_entity_ids 里拿到那个唯一的 entity_id
        target_entity_id = distinct_entity_ids.first()  # 可能为 None
        if target_entity_id is None:
            return Response({"error": "条目中没有绑定测评对象"}, status=400)
        # 再去查一次拿到它的 name
        try:
            entity_obj = AiEvaluationEntity.objects.get(pk=target_entity_id)
        except AiEvaluationEntity.DoesNotExist:
            return Response({"error": "测评对象不存在"}, status=404)
        target_entity_name = entity_obj.name
        # 4. 创建历史记录将 ID 写入 dataset/agent/entity 字段
        ts = timezone.now().strftime('%Y%m%d%H%M')
        record = AiEvaluationRecord.objects.create(name=f"{target_entity_name}-{dataset.name}-{ts}",dataset=str(dataset_id),agent=",".join(str(aid) for aid in agent),entity=str(target_entity_id), )
        # 5. 启动后台线程
        worker = threading.Thread(target=self._run_evaluate_by_agents,args=(record.id, dataset_id, agent),daemon=True)
        worker.start()

        return Response({"message": "评估已开始，可在历史记录中查看结果", "record_id": record.id},status=200)
    def _run_evaluate_by_agents(self, record_id, dataset_id, agent_id):
        """
        后台线程：对 dataset_id 下每个条目，逐个 agent 调用 AIAgentServer 去评分，
        并把结果存入 AiEvaluationRecordEntry。
        """
        try:
            record = AiEvaluationRecord.objects.get(pk=record_id)
            entries = AiEvaluationDatasetEntry.objects.filter(dataset_id=dataset_id)
            for entry in entries:
                # 每条 entry 的结果字典，先放 question
                single = {"question": entry.question}
                for aid in agent_id:
                    agent_obj = AiServerAgent.objects.get(pk=aid)
                    cfg = AiServerAgentSerializer(agent_obj).data
                    # 构造 user_content
                    parts = [
                        f"Question: {entry.question}",
                        f"Ground_truth: {entry.ground_truth or ''}",
                        *([f"Contexts: {entry.contexts}"] if entry.contexts else []),
                        f"Answer: {entry.answer or ''}",
                    ]
                    user_content = "\n".join(parts)
                    # 调用 AIAgentServer
                    try:
                        raw = AIAgentServer(agent_config=cfg,user_content=user_content).agent_server().strip()
                    except Exception as e:
                        logger.error(f"[Agent {agent_obj.name}] 调用失败：{e}")
                        raw = "0"
                    # 解析分数
                    score = 0.0
                    try:
                        score = round(float(raw), 2)
                    except ValueError:
                        # 支持 JSON 格式或正则提取
                        try:
                            j = json.loads(raw)
                            score = round(float(j.get("score", 0)), 2)
                        except Exception:
                            m = re.search(r'"score"\s*:\s*(\d+(\.\d+)?)', raw)
                            if m:
                                score = round(float(m.group(1)), 2)
                    single[agent_obj.name] = score
                # 写入这一条 entry 的所有 agent 结果（JSON 存入 result 字段）
                AiEvaluationRecordEntry.objects.create(record=record,result=json.dumps(single, ensure_ascii=False))
            logger.info(f"异步评测完成 record_id={record_id}")
        except Exception as ex:
            logger.exception(f"异步评测出错 record_id={record_id}：{ex}")

class AiEvaluationFileView(APIView):
    def post(self, request):
        return pagination_query(cls=AiEvaluationFile, cls_serializer=AiEvaluationFileSerializer, request=request)


class AiEvaluationFileDatasetView(APIView):
    parser_classes = (JSONParser,)

    def post(self, request):
        # 前端传 dataset_id，就只查该数据集下挂载的文件
        dataset_id = request.data.get('dataset_id')
        if dataset_id is not None:
            try:
                ds = AiEvaluationDataset.objects.get(pk=int(dataset_id))
            except (ValueError, AiEvaluationDataset.DoesNotExist):
                return Response({"error": "数据集不存在"}, status=status.HTTP_404_NOT_FOUND)

            # 如果数据集有 file 外键，就返回它；否则返回空列表
            if ds.file_id:
                qs = AiEvaluationFile.objects.filter(pk=ds.file_id)
            else:
                qs = AiEvaluationFile.objects.none()

            serializer = AiEvaluationFileSerializer(qs, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        # 否则走原来的分页查询，返回所有文件
        return pagination_query(cls=AiEvaluationFile, cls_serializer=AiEvaluationFileSerializer, request=request)


class AiEvaluationFileInfo(APIView):
    def get(self, request):
        return get_by_id(cls=AiEvaluationFile, cls_serializer=AiEvaluationFileSerializer, request=request)

    def post(self, request):
        return post_by_id(cls=AiEvaluationFile, cls_serializer=AiEvaluationFileSerializer, request=request)

    def delete(self, request):
        return delete_by_id(cls=AiEvaluationFile, request=request)


class AiEvaluationDatasetTree(APIView):
    def post(self, request):
        return tree_query(cls=AiEvaluationDataset, cls_serializer=AiEvaluationDatasetSerializer)

    def delete(self, request):
        #     """
        #     DELETE /ai_evaluation/dataset/tree
        #     {"dataset_id": 5}
        # 1. 拿到 dataset_id（支持 body 或 query）
        dataset_id = request.data.get('dataset_id') or request.query_params.get('dataset_id')
        if not dataset_id:
            return api_response(status=400, message="缺少参数 dataset_id")

        # 2. 校验节点是否存在
        try:
            node = AiEvaluationDataset.objects.get(pk=int(dataset_id))
        except (ValueError, AiEvaluationDataset.DoesNotExist):
            return api_response(status=404, message=f"节点 ID={dataset_id} 不存在")

        # 3. 如果存在子节点，则拒绝删除
        if node.children.exists():
            return api_response(status=400, message="当前节点存在子节点，无法删除")

        # 4. 真正删除
        node.delete()
        return api_response(status=200, message="节点删除成功")


class AiEvaluationDatasetInfo(APIView):
    def post(self, request):
        return post_by_id(cls=AiEvaluationDataset, cls_serializer=AiEvaluationDatasetSerializer, request=request)


class AiEvaluationDatasetTagView(APIView):
    def post(self, request):
        return pagination_query(cls=AiEvaluationDatasetTag, cls_serializer=AiEvaluationDatasetTagSerializer,
                                request=request)


class AiEvaluationDatasetTagInfo(APIView):
    def get(self, request):
        return get_by_id(cls=AiEvaluationDatasetTag, cls_serializer=AiEvaluationDatasetTagSerializer, request=request)

    def post(self, request):
        return post_by_id(cls=AiEvaluationDatasetTag, cls_serializer=AiEvaluationDatasetTagSerializer, request=request)

    def delete(self, request):
        return delete_by_id(cls=AiEvaluationDatasetTag, request=request)


class AiEvaluationEntityView(APIView):
    parser_classes = (JSONParser,)

    def get(self, request):
        # 支持 GET /…/evaluation/object/list
        return pagination_query(cls=AiEvaluationEntity, cls_serializer=AiEvaluationEntitySerializer, request=request)

    def post(self, request):
        # 支持 POST /…/evaluation/object/list
        return pagination_query(cls=AiEvaluationEntity, cls_serializer=AiEvaluationEntitySerializer, request=request)


class AiEvaluationEntityInfo(APIView):
    def get(self, request):
        return get_by_id(cls=AiEvaluationEntity, cls_serializer=AiEvaluationEntitySerializer, request=request)

    def post(self, request):
        return post_by_id(cls=AiEvaluationEntity, cls_serializer=AiEvaluationEntitySerializer, request=request)

    def delete(self, request):
        return delete_by_id(cls=AiEvaluationEntity, request=request)


class AiEvaluationDatasetEntryView(APIView):
    def post(self, request):
        return pagination_query(cls=AiEvaluationDatasetEntry, cls_serializer=AiEvaluationDatasetEntrySerializer,
                                request=request)


class AiEvaluationDatasetEntryInfo(APIView):
    def get(self, request):
        return get_by_id(cls=AiEvaluationDatasetEntry, cls_serializer=AiEvaluationDatasetEntrySerializer,
                         request=request)

    def post(self, request):
        return post_by_id(cls=AiEvaluationDatasetEntry, cls_serializer=AiEvaluationDatasetEntrySerializer,
                          request=request)

    def delete(self, request):
        return delete_by_id(cls=AiEvaluationDatasetEntry, request=request)


# 历史记录
class AiEvaluationRecordView(APIView):
    # def post(self, request):
    #     return pagination_query(cls=AiEvaluationRecord, cls_serializer=AiEvaluationRecordSerializer,request=request)
    def post(self, request):
        """
        POST /ai_evaluation/evaluation/history/list
        支持用 name/dataset/agent/entity 做模糊过滤，
        并走公共 pagination_query。
        """
        # 1. 构造过滤规则
        page_rule = []
        if request.data.get('name'):
            page_rule.append({'field': 'name', 'rule': 'contains', 'value': request.data['name']})
        if request.data.get('dataset'):
            page_rule.append({'field': 'dataset', 'rule': 'contains', 'value': request.data['dataset']})
        if request.data.get('agent'):
            page_rule.append({'field': 'agent', 'rule': 'contains', 'value': request.data['agent']})
        if request.data.get('entity'):
            page_rule.append({'field': 'entity', 'rule': 'contains', 'value': request.data['entity']})

        # 2. 注入到请求体
        request.data['pageRule'] = page_rule

        # 3. 调用公共分页查询
        return pagination_query(cls=AiEvaluationRecord, cls_serializer=AiEvaluationRecordSerializer, request=request)


class AiEvaluationRecordInfo(APIView):
    def get(self, request):
        return get_by_id(cls=AiEvaluationRecord, cls_serializer=AiEvaluationRecordSerializer, request=request)

    def post(self, request):
        return post_by_id(cls=AiEvaluationRecord, cls_serializer=AiEvaluationRecordSerializer, request=request)

    def delete(self, request):
        return delete_by_id(cls=AiEvaluationRecord, request=request)


# 历史测评对象记录
class AiEvaluationRecordEntryView(APIView):
    def post(self, request):
        """
        POST /ai_evaluation/evaluation/history/detail/list
        不开启分页查询
        {
          "pageEnable": false,
          "pageRule": [
        {
           "field": "record",
           "rule":  "is",
           "value": 28
        }
       ]
       }
       开启分页查询
       {
         "pageEnable": true,
         "pageNum": 1,
         "pageSize": 20,
        "pageRule": [
        {
          "field": "record",
          "rule":  "is",
          "value": 28
         }
        ]
       }
        """
        record_id = request.data.get('record_id')
        if record_id is not None:
            request.data.setdefault('pageRule', [])
            request.data['pageRule'].append({'field': 'record','rule': 'is','value': record_id,})

        print("DEBUG pageRule:", request.data.get('pageRule'))

        return pagination_query(
            cls=AiEvaluationRecordEntry,
            cls_serializer=AiEvaluationRecordEntrySerializer,
            request=request)


class AiEvaluationRecordEntryInfo(APIView):
    def get(self, request):
        return get_by_id(cls=AiEvaluationRecordEntry, cls_serializer=AiEvaluationRecordEntrySerializer, request=request)

    def post(self, request):
        return post_by_id(cls=AiEvaluationRecordEntry, cls_serializer=AiEvaluationRecordEntrySerializer,
                          request=request)

    def delete(self, request):
        return delete_by_id(cls=AiEvaluationRecordEntry, request=request)


# Excel上传
class ExcelImportView(APIView):
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    @transaction.atomic()
    def post(self, request):
        # 1. 验证 dataset_id
        dataset_id = request.data.get('dataset_id')
        try:
            dataset_obj = AiEvaluationDataset.objects.get(id=dataset_id)
        except AiEvaluationDataset.DoesNotExist:
            return JsonResponse({"error": f"数据集 ID={dataset_id} 不存在，无法导入。"}, status=400)
        # 2. 读取上传的 .xlsx
        f = request.FILES.get('file')
        if not f or not f.name.lower().endswith('.xlsx'):
            return JsonResponse({"error": "只支持 .xlsx 文件导入。"}, status=400)
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        # 3. 解析表头，找到每列的下标
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header_lower = [str(c).strip().lower() for c in header]
        def find_idx(*names):
            for name in names:
                nl = name.lower()
                if nl in header_lower:
                    return header_lower.index(nl)
            return None
        question_idx = find_idx('question', '标准问题')
        ground_truth_idx = find_idx('standard answer', '标准答案')
        answer_idx = find_idx('ai answer', '测评答案')
        contexts_idx = find_idx('contexts', '上下文')
        update_time_idx = find_idx('updated', '更新时间')
        tag_idx = find_idx('tag', '问题标签')
        entries = []
        skipped = 0
        # 4. 从表头下一行开始循环真实数据
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 整行全空就跳过
            if not any(cell is not None for cell in row):
                skipped += 1
                continue
            # 必填列校验：只校验 question
            q = row[question_idx]
            if not q:
                skipped += 1
                continue
            # 其他列
            g = row[ground_truth_idx] if ground_truth_idx is not None else ''
            a = row[answer_idx] if answer_idx is not None else ''
            ctx = row[contexts_idx] if contexts_idx is not None else ''
            ut = row[update_time_idx] if update_time_idx is not None else None
            tg = row[tag_idx] if tag_idx is not None else None
            entries.append(
                {"question": str(q).strip(), "ground_truth": g or "", "answer": a or "", "contexts": ctx or "","update_time": ut, "tag": tg, })
        # 5. 缓存 tag 和 entity
        tag_cache = {t.name: t for t in AiEvaluationDatasetTag.objects.all()}
        # 6. 加载现有记录并按 question 建立映射
        existing_q_map = {
            entry.question: entry
            for entry in AiEvaluationDatasetEntry.objects.filter(dataset=dataset_obj)
        }
        created_count = 0
        updated_count = 0
        for rec in entries:
            # 处理 contexts 列
            ctx_list = rec["contexts"].split(';') if rec["contexts"] else []
            # 解析时间
            parsed_dt = None
            if rec["update_time"]:
                for fmt in ("%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M:%S"):
                    try:
                        parsed_dt = datetime.datetime.strptime(str(rec["update_time"]), fmt)
                        break
                    except ValueError:
                        continue
            # 构造参数字典
            params = {"ground_truth": rec["ground_truth"], "answer": rec["answer"],"contexts": json.dumps(ctx_list, ensure_ascii=False), "tag": tag_cache.get(rec["tag"]), }
            if parsed_dt:
                params["update_time"] = parsed_dt
            if rec["question"] in existing_q_map:
                # 已存在→更新
                entry = existing_q_map[rec["question"]]
                for attr, val in params.items():
                    setattr(entry, attr, val)
                entry.save()
                updated_count += 1
            else:
                # 不存在就新建
                AiEvaluationDatasetEntry.objects.create(question=rec["question"],dataset=dataset_obj,**params)
                created_count += 1
        # 7. 返回结果
        return JsonResponse({
            "status": "success","message": (f"导入完成：{updated_count} 条更新，"f"{created_count} 条新建，{skipped} 条跳过。")})


# 一键更新
class DatasetEvaluateView(APIView):
    """
    POST /ai_evaluation/dataset/evaluate
    { entity_id:1}
    { entity_id:1,dataset_id:1}
    调试模式：仅传 entity_id（可在 query_params 或 body），后端同步使用 "Hello, world!" 测试并返回。
    批量模式：传 dataset_id 和 entity_id，后台线程异步调用，立即返回格式化消息（并在后台更新 answer、update_time、entity）。
    """

    def post(self, request):
        # 支持从 JSON body 或 URL query params 获取参数
        dataset_id = request.data.get('dataset_id') or request.query_params.get('dataset_id')
        entity_id = request.data.get('entity_id') or request.query_params.get('entity_id')

        # 转换为整数
        try:
            dataset_id = int(dataset_id) if dataset_id is not None else None
        except (TypeError, ValueError):
            dataset_id = None
        try:
            entity_id = int(entity_id) if entity_id is not None else None
        except (TypeError, ValueError):
            entity_id = None

        # 调试模式：只有 entity_id
        if entity_id and not dataset_id:
            try:
                entity = AiEvaluationEntity.objects.get(id=entity_id)
            except AiEvaluationEntity.DoesNotExist:
                return api_response(status=400, message='评测对象 ID 不存在')

            test_query = "Hello, world!"
            cfg = entity.get_request_config()
            cfg['body'] = self._build_body_with_exec(cfg.get('body', {}), test_query)

            resp = GenericAIService.get_answer(cfg)
            if resp.get('answer'):
                return api_response(status=200, data={'answer': resp['answer']}, message='调试评测完成')
            return api_response(status=500, message=resp.get('error', '未知错误'))

        # 批量模式：同时传 dataset_id 和 entity_id
        if dataset_id and entity_id:
            try:
                entity = AiEvaluationEntity.objects.get(id=entity_id)
            except AiEvaluationEntity.DoesNotExist:
                return api_response(status=400, message='评测对象 ID 不存在')
            worker = threading.Thread(target=self._run_batch_evaluation, args=(dataset_id, entity))
            worker.daemon = True
            worker.start()
            return api_response(status=200, message='更新开始')
        # 参数不足
        return api_response(status=400, message='缺少 dataset_id 或 entity_id')

    def _run_batch_evaluation(self, dataset_id, entity):
        try:
            dataset = AiEvaluationDataset.objects.get(id=dataset_id)
        except AiEvaluationDataset.DoesNotExist:
            return

        qs_all = AiEvaluationDatasetEntry.objects.filter(dataset=dataset)

        for entry in qs_all.iterator():
            # 重置旧数据
            entry.answer = ''
            entry.update_time = None
            entry.entity = entity  # 更新为指定评测对象
            entry.save(update_fields=['answer', 'update_time', 'entity'])

            # 原始问题
            q = entry.question
            cfg = entity.get_request_config()
            cfg['body'] = self._build_body_with_exec(cfg.get('body', {}), q)

            # 调用第三方服务
            resp = GenericAIService.get_answer(cfg)
            if resp.get('answer'):
                entry.answer = resp['answer']
                entry.status = 'completed'
                entry.update_time = timezone.now().replace(microsecond=0)
                entry.save(update_fields=['answer', 'status', 'update_time'])
            else:
                entry.status = 'false'
                entry.save(update_fields=['status'])

    def _build_body_with_exec(self, original_body: dict, query: str) -> dict:
        """
        用 exec+repr 的方式，在 Python 字面量层面替换 '{{query}}'。
        """
        body = copy.deepcopy(original_body)
        body_repr = repr(body)
        body_repr = body_repr.replace("'{{query}}'", repr(query))
        namespace = {}
        exec(f"result = {body_repr}", {}, namespace)
        return namespace['result']


#  文件上传与下载
class DatasetFileUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        # 从请求体中读取 dataset_id
        dataset_id = request.data.get('dataset_id')
        if not dataset_id:
            return Response({"error": "缺少 dataset_id"}, status=400)

        # 调用公共函数处理上传
        return dataset_file_upload(request=request, dataset_model=AiEvaluationDataset, dataset_id=dataset_id, )


class DatasetFileDownloadView(APIView):
    def get(self, request):
        # 从查询参数中读取 dataset_id
        dataset_id = request.GET.get('dataset_id')
        if not dataset_id:
            return Response({"error": "缺少 dataset_id"}, status=400)

        # 调用公共函数处理下载
        return dataset_file_download(dataset_model=AiEvaluationDataset, dataset_id=dataset_id, )
